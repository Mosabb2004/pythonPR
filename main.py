from tkinter import *
from tkinter import ttk
import datetime
import openpyxl 
from openpyxl import Workbook


root  =Tk()
root.geometry('960x552')
root.iconbitmap('img/icon.ico')
root.title('shop market')

now = datetime.datetime.now()
date=now.strftime("%Y-%m-%d")


#---------------------fiyatlar------------------------

menu ={
    0:['testere',200],
    1:['el arabasaı',500],
    2:['kazma',250],
    3:['kürek',200],
    4:['çekiç',150],
    5:['kova',100],
    6:['inşaat kaskı',175],
    7:['maket bıçağı',50],
    8:['segman pense',90],
    9:['pense',75],
    10:['tornavida',50],
    11:['dril',600],
}

def bill():
    global En_name
    global En_phone
    global En_adres
    global En_tarih
    global En_toplam

    root.geometry('1210x552')
    F3 =Frame(root,bg='#5F7161',width=250,height=550,bd=2,relief=GROOVE)
    F3.place(x=955,y=1)

    L_name=Label(F3,text='müşteri adı',bg='#5F7161',fg='white')
    L_name.place(x=10,y=10)
    En_name=Entry(F3,width=24,font=('Tajawal',12),justify=CENTER)
    En_name.place(x=13,y=40)

    L_phone=Label(F3,text='müşteri numarası',bg='#5F7161',fg='white')
    L_phone.place(x=10,y=75)
    En_phone=Entry(F3,width=24,font=('Tajawal',12),justify=CENTER)
    En_phone.place(x=12,y=105)

    L_adres=Label(F3,text='müşteri adresi',bg='#5F7161',fg='white')
    L_adres.place(x=12,y=140)
    En_adres=Entry(F3,width=24,font=('Tajawal',12),justify=CENTER)
    En_adres.place(x=12,y=170)

    L_toplam=Label(F3,text='toplam',bg='#5F7161',fg='white')
    L_toplam.place(x=12,y=205)
    En_toplam=Entry(F3,width=24,font=('Tajawal',12),justify=CENTER)
    En_toplam.place(x=12,y=235)

    L_tarih=Label(F3,text='tarih',bg='#5F7161',fg='white')
    L_tarih.place(x=12,y=270)
    En_tarih=Entry(F3,width=24,font=('Tajawal',12),justify=CENTER)
    En_tarih.place(x=12,y=300)


    add_button=Button(F3,text='kaydet',width=30,cursor='hand2',bg='#EDDBC0' ,command=save)
    add_button.place(x=12,y=350)

    add_button=Button(F3,text='yenile',width=30,cursor='hand2',bg='#EDDBC0',command=temizle2)
    add_button.place(x=12,y=390)

    add_button=Button(F3,text='müşteri ara',width=30,cursor='hand2',bg='#EDDBC0')
    add_button.place(x=12,y=430)

    add_button=Button(F3,text='faturayı sil',width=30,cursor='hand2',bg='#EDDBC0',command=temizle1)
    add_button.place(x=12,y=470)

    total=0
    for item in trv.get_children():
        trv.delete(item)

    for i in range(len(sb)):
        if(int(sb[i].get())>0):
            price=int(sb[i].get())*menu[i][1]
            total=total+price
            myst=(str(menu[i][1]),str(sb[i].get()),str(price))
            trv.insert("",'end',iid=i,text=menu[i][0],values=myst)
    finall=total
    En_toplam.insert('1',str(finall)+'$')
    En_tarih.insert('1',str(date))


def temizle():
    for item in trv.get_children():
        trv.delete(item)
    En_name.delete('0',END)
    En_phone.delete('0',END)
    En_adres.delete('0',END)
    En_tarih.delete('0',END)
    En_toplam.delete('0',END)


def temizle1():
    En_name.delete('0',END)
    En_phone.delete('0',END)
    En_adres.delete('0',END)
    En_tarih.delete('0',END)
    En_toplam.delete('0',END)    


def temizle2():
    En_name.delete(0,END)
    En_phone.delete(0,END)
    En_adres.delete(0,END)
    En_tarih.delete(0,END)
    En_toplam.delete(0,END) 

    for item in trv.get_children():
        trv.delete(item) 

    for sb_item in sb:
        sb_item.delete(0,END)
        sb_item.insert(0,'0')


def kapat():
    root.quit()    




wb=Workbook()
ws=wb.active

ws.title='müşteri'
ws["A1"]='Ad'
ws["B1"]='telefon numarası'
ws["C1"]='adres'
ws["D1"]='toplam'
ws["E1"]='tarih'
wb.save('Arel.xlsx')

def save():
    name=En_name.get()
    phone=En_phone.get()
    adres=En_adres.get()
    toplam=En_toplam.get()
    tarih=En_tarih.get()

    excel =openpyxl.load_workbook('Arel.xlsx')
    file = excel.active

    row =file.max_row +1
 
    file.cell(column=1,row=row,value=name)
    file.cell(column=2,row=row,value=phone)
    file.cell(column=3,row=row,value=adres)
    file.cell(column=4,row=row,value=toplam)
    file.cell(column=5,row=row,value=tarih)

    excel.save('Arel.xlsx')



F1 = Frame(root,bg='silver',width=600,height=550)
F1.place(x=1,y=1)


img_menu1=PhotoImage(file='img/1.png')
img_menu2=PhotoImage(file='img/2.png')
img_menu3=PhotoImage(file='img/3.png')
img_menu4=PhotoImage(file='img/4.png')
img_menu5=PhotoImage(file='img/5.png')
img_menu6=PhotoImage(file='img/6.png')
img_menu7=PhotoImage(file='img/7.png')
img_menu8=PhotoImage(file='img/8.png')
img_menu9=PhotoImage(file='img/9.png')
img_menu10=PhotoImage(file='img/10.png')
img_menu11=PhotoImage(file='img/11.png')
img_menu12=PhotoImage(file='img/12.png')

title =Label(F1,text='inşaat melzemeleri',font=('Tajawal 13'),fg='white',bg='#5F7161',width=70)
title.place(x=0,y=0)

menu1=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu1,text='testere',compound=TOP)
menu1.place(x=30,y=45)

menu2=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu2,text='el arabası',compound=TOP)
menu2.place(x=170,y=45)

menu3=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu3,text='kazma',compound=TOP)
menu3.place(x=310,y=45)

menu4=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu4,text='kürek',compound=TOP)
menu4.place(x=450,y=45)

menu5=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu5,text='çekiç',compound=TOP)
menu5.place(x=30,y=180)

menu6=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu6,text='kova',compound=TOP)
menu6.place(x=170,y=180)

menu7=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu7,text='inşaat kaskı',compound=TOP)
menu7.place(x=310,y=180)

menu8=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu8,text='maket bıçağı',compound=TOP)
menu8.place(x=450,y=180)

menu9=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu9,text='segman pense',compound=TOP)
menu9.place(x=30,y=325)

menu10=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu10,text='pense',compound=TOP)
menu10.place(x=170,y=325)

menu11=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu11,text='tornavida',compound=TOP)
menu11.place(x=310,y=325)

menu12=Button(F1,width=88,bg='#EFEAD8',bd=1,relief=SOLID,cursor='hand2',height=85,image=img_menu12,text='dril',compound=TOP)
menu12.place(x=450,y=325)



sb = []
font1 =('Times',12,'normal')

sv1=IntVar()
sv2=IntVar()
sv3=IntVar()
sv4=IntVar()
sv5=IntVar()
sv6=IntVar()
sv7=IntVar()
sv8=IntVar()
sv9=IntVar()
sv10=IntVar()
sv11=IntVar()
sv12=IntVar()

sb1 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv1)
sb1.place(x=30,y=140)
sb.append(sb1)

sb2 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv2)
sb2.place(x=170,y=140)
sb.append(sb2)

sb3 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv3)
sb3.place(x=310,y=140)
sb.append(sb3)

sb4 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv4)
sb4.place(x=450,y=140)
sb.append(sb4)

sb5 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv5)
sb5.place(x=30,y=275)
sb.append(sb5)

sb6 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv6)
sb6.place(x=170,y=275)
sb.append(sb6)

sb7 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv7)
sb7.place(x=310,y=275)
sb.append(sb7)

sb8 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv8)
sb8.place(x=450,y=275)
sb.append(sb8)

sb9 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv9)
sb9.place(x=30,y=420)
sb.append(sb9)

sb10 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv10)
sb10.place(x=170,y=420)
sb.append(sb10)

sb11 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv11)
sb11.place(x=310,y=420)
sb.append(sb11)

sb12 = Spinbox(F1,from_=0,to_=5,font=font1,width=10,textvariable=sv12)
sb12.place(x=450,y=420)
sb.append(sb12)


b1=Button(F1,text='AL',fg='white',font=('Tajawal 12'),bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,width=12,command=bill)
b1.place(x=30,y=500)

b2=Button(F1,text='YENİ FATURA',fg='white',font=('Tajawal 12'),bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,width=12,command=temizle)
b2.place(x=160,y=500)

b3=Button(F1,text='KIRALAMA',fg='white',font=('Tajawal 12'),bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,width=12)
b3.place(x=290,y=500)

b4=Button(F1,text='KAPAT',fg='white',font=('Tajawal 12'),bg='#6D8B74',bd=1,relief=SOLID,cursor='hand2',height=1,width=12,command=kapat)
b4.place(x=420,y=500)



#---------------------------FRAME2------------------------

F2=Frame(root,bg='gray',width=350,height=550)
F2.place(x=604,y=1)

trv = ttk.Treeview(F2,selectmode='browse')
trv.place(x=1,y=1,width=347,height=550)

trv["columns"]=('1','2','3')
trv.column("#0",width=80,anchor='c')
trv.column("1",width=50,anchor='c')
trv.column("2",width=45,anchor='c')
trv.column("3",width=60,anchor='c')
trv.heading("#0",text='ürün',anchor='c')
trv.heading("1",text='fiyat',anchor='c')
trv.heading("2",text='addet',anchor='c')
trv.heading("3",text='toplam',anchor='c')



root.mainloop()
